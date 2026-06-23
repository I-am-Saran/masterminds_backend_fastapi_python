# REFERENCE_ONLY_MODULE — legacy Kaizen module; preserved for compatibility.
from fastapi import APIRouter, HTTPException, Header, Query
import json
from typing import Optional, List, Dict, Any
from services.db_service import execute_query, local_db
from services.auth_service import get_user_from_token, auth_guard
from services.rbac_service import get_user_accessible_projects, get_user_accessible_project_names
from utils.error_handler import handle_api_error, handle_endpoint_error
from services.rbac_service import require_permission
from datetime import datetime, date, timedelta

OPEN_STATUSES = [
    "OPEN",
    "IN_PROGRESS",
    "RESOLVED",
    "READY FOR DEPLOYMENT",
    "READY FOR QA",
    "QA IN PROGRESS",
    "QA REOPENED",
    "NEED CLARIFICATION FROM DATA TEAM",
    "NEED CLARIFICATION FROM PRODUCT TEAM",
    "NEED CLARIFICATION FROM DEV TEAM",
    "NEED CLARIFICATION FROM QA TEAM",
    "NEED CLARIFICATION FROM DEVOPS",
    "NEED CLARIFICATION FROM COMPLIANCE",
    "NEED CLARIFICATION FROM SECURITY",
    "NEED CLARIFICATION FROM INFRA",
    "NEED CLARIFICATION FROM VENDOR",
    "REVIEW IN PROGRESS"
]

QA_PASSED_STATUSES = ["QA PASSED"]

SNAPSHOT_OPEN_SUPER_STATUSES = (
    "NEW",
    "ACTIVE_DEV",
    "REVIEW",
    "READY",
    "REOPENED",
    "WAITING",
    "RESOLVED_PENDING",
)

SNAPSHOT_CLOSED_SUPER_STATUSES = (
    "DELIVERED",
    "ACCEPTED",
    "CLOSED",
    "REJECTED",
    "DEFERRED",
    "DUPLICATE",
)

# Lean column list for QA dashboard — excludes Attachments, Comments, ActivityLog, etc.
QA_BUG_SELECT_COLUMNS = [
    '"Bug ID"',
    '"Project"',
    '"Product"',
    '"Status"',
    '"Priority"',
    '"Severity"',
    '"Component"',
    '"created_at"',
    '"updated_at"',
    '"Changed"',
    '"Bug Age (in days)"',
    '"Assignee"',
    '"Assignee Real Name"',
    '"Reporter"',
    '"Reporter Real Name"',
    '"Project Owner"',
    '"Project Owner Name"',
    '"Summary"',
    '"tester_type"',
    '"Deadline"',
    '"Defect type"',
]

# Heavy JSON / blob columns intentionally omitted from dashboard fetches:
# "Attachments", "Comments", "Comment", "ActivityLog", "Description", "Steps to Reproduce"


def ensure_bug_indexes():
    """Best-effort index creation (dev/local). Production: run scripts/sql/qa_dashboard_production_indexes.sql"""
    index_statements = [
        'CREATE INDEX IF NOT EXISTS idx_bugs_tenant_deleted ON "bugs" ("tenant_id") WHERE "is_deleted" IS DISTINCT FROM TRUE',
        'CREATE INDEX IF NOT EXISTS idx_bugs_tenant_created ON "bugs" ("tenant_id", "created_at") WHERE "is_deleted" IS DISTINCT FROM TRUE',
        'CREATE INDEX IF NOT EXISTS idx_bugs_tenant_status ON "bugs" ("tenant_id", "Status") WHERE "is_deleted" IS DISTINCT FROM TRUE',
        'CREATE INDEX IF NOT EXISTS idx_bugs_tenant_product ON "bugs" ("tenant_id", "Product") WHERE "is_deleted" IS DISTINCT FROM TRUE',
        'CREATE INDEX IF NOT EXISTS idx_bugs_tenant_project ON "bugs" ("tenant_id", "Project") WHERE "is_deleted" IS DISTINCT FROM TRUE',
        'CREATE INDEX IF NOT EXISTS idx_bugs_tenant_project_created ON "bugs" ("tenant_id", "Project", "created_at") WHERE "is_deleted" IS DISTINCT FROM TRUE',
        'CREATE INDEX IF NOT EXISTS idx_bugs_tenant_changed ON "bugs" ("tenant_id", "Changed") WHERE "is_deleted" IS DISTINCT FROM TRUE',
        'CREATE INDEX IF NOT EXISTS idx_qa_bug_snapshot_daily_date_project ON qa_bug_snapshot_daily (snapshot_date, project_name)',
    ]
    try:
        for stmt in index_statements:
            execute_query(stmt, fetch_all=False)
    except Exception:
        pass

# Common function to parse dates
def parse_date(date_val: Any) -> Optional[date]:
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val
    
    # It is a string
    try:
        date_str = str(date_val)
        # Handle 'T' or space separator
        clean_str = date_str.replace('T', ' ').split(' ')[0]
        return datetime.strptime(clean_str, "%Y-%m-%d").date()
    except (ValueError, AttributeError):
        return None

def safe_sort_datetime(date_val: Any) -> datetime:
    """Safely converts any date/time value into a naive datetime object for sorting.
    Always returns a datetime object (datetime.min for invalid/null values).
    """
    if not date_val:
        return datetime.min
    if isinstance(date_val, datetime):
        return date_val.replace(tzinfo=None)
    if isinstance(date_val, date):
        return datetime.combine(date_val, datetime.min.time())
    
    try:
        date_str = str(date_val).strip()
        if 'T' in date_str or ' ' in date_str:
            clean_str = date_str.split('.')[0].replace('Z', '')
            try:
                return datetime.fromisoformat(clean_str).replace(tzinfo=None)
            except ValueError:
                pass
        
        clean_str = date_str.replace('T', ' ').split(' ')[0]
        return datetime.strptime(clean_str, "%Y-%m-%d")
    except (ValueError, AttributeError, TypeError):
        return datetime.min

def normalize_status(status: Any) -> str:
    return str(status or "").strip().upper()


def _bug_super_status(status: Any) -> str:
    """Map raw bug Status to qa_bug_snapshot_daily.super_status."""
    s = normalize_status(status)
    if s == "OPEN":
        return "NEW"
    if s in ("IN_PROGRESS", "DEV IN PROGRESS"):
        return "ACTIVE_DEV"
    if s == "REVIEW IN PROGRESS":
        return "REVIEW"
    if s == "READY FOR DEPLOYMENT":
        return "READY"
    if s in ("QA REOPENED", "UAT REOPENED"):
        return "REOPENED"
    if "NEED CLARIFICATION" in s:
        return "WAITING"
    if s == "RESOLVED":
        return "RESOLVED_PENDING"
    if s == "DEPLOYED TO PROD":
        return "DELIVERED"
    if s in ("QA PASSED", "UAT PASSED"):
        return "ACCEPTED"
    if s == "CLOSED":
        return "CLOSED"
    if s in ("REJECTED", "INVALID"):
        return "REJECTED"
    if s in ("DESCOPED", "DEFERRED"):
        return "DEFERRED"
    if s == "DUPLICATE":
        return "DUPLICATE"
    return "OTHER"


def _bug_age_bucket(created_at: Any) -> str:
    created = parse_date(created_at)
    if not created:
        return "30+ DAYS"
    age = (date.today() - created).days
    if age <= 2:
        return "0-2 DAYS"
    if age <= 7:
        return "3-7 DAYS"
    if age <= 14:
        return "8-14 DAYS"
    if age <= 30:
        return "15-30 DAYS"
    return "30+ DAYS"


def _priority_upper(priority: Any) -> str:
    return str(priority or "UNKNOWN").strip().upper()


def _aggregate_bugs_snapshot_metrics(bugs: List[dict]) -> Dict[str, int]:
    """Build one-day trend metrics from live bugs (fallback when snapshot table is empty)."""
    metrics = {
        "open_bugs": 0,
        "closed_bugs": 0,
        "total_bugs": 0,
        "blocker_high": 0,
        "sla_breached": 0,
        "closed_blocker": 0,
        "closed_high": 0,
        "closed_medium": 0,
        "closed_low": 0,
    }
    open_set = set(SNAPSHOT_OPEN_SUPER_STATUSES)
    closed_set = set(SNAPSHOT_CLOSED_SUPER_STATUSES)

    for b in bugs or []:
        super_status = _bug_super_status(b.get("Status"))
        pri = _priority_upper(b.get("Priority"))
        age_bucket = _bug_age_bucket(b.get("created_at"))
        metrics["total_bugs"] += 1

        if super_status in open_set:
            metrics["open_bugs"] += 1
            if pri in ("BLOCKER", "HIGH", "P1", "P2", "CRITICAL", "MAJOR"):
                metrics["blocker_high"] += 1
            if age_bucket == "30+ DAYS":
                metrics["sla_breached"] += 1
        elif super_status in closed_set:
            metrics["closed_bugs"] += 1
            if pri in ("BLOCKER", "P1", "CRITICAL"):
                metrics["closed_blocker"] += 1
            elif pri in ("HIGH", "P2", "MAJOR"):
                metrics["closed_high"] += 1
            elif pri in ("MEDIUM", "P3"):
                metrics["closed_medium"] += 1
            elif pri in ("LOW", "P4"):
                metrics["closed_low"] += 1

    return metrics


def _snapshot_metrics_to_chart_point(d_str: str, row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "date": d_str,
        "open": int(row.get("open_bugs") or 0),
        "closed": int(row.get("closed_bugs") or 0),
        "total": int(row.get("total_bugs") or 0),
        "blockerHigh": int(row.get("blocker_high") or 0),
        "slaBreached": int(row.get("sla_breached") or 0),
        "closedBlocker": int(row.get("closed_blocker") or 0),
        "closedHigh": int(row.get("closed_high") or 0),
        "closedMedium": int(row.get("closed_medium") or 0),
        "closedLow": int(row.get("closed_low") or 0),
    }


router = APIRouter(prefix="/qa-dashboard", tags=["qa-dashboard"])


def get_user_map(tenant_id: str) -> Dict[str, Dict[str, str]]:
    """Fetch all users and return a mapping of id -> {full_name, email} for a tenant"""
    try:
        query = 'SELECT id, full_name, email FROM "users" WHERE "tenant_id" = %s'
        rows = execute_query(query, (tenant_id,), fetch_all=True) or []
        # Convert UUID to str for key
        return {str(u['id']): {"name": u['full_name'], "email": u['email']} for u in rows}
    except:
        return {}

def get_valid_project_names(tenant_id: str) -> set:
    """Fetch all valid project names (application_name) for a tenant."""
    try:
        query = local_db.table("projects").select("application_name")
        if tenant_id:
            query = query.eq("tenant_id", tenant_id)
        resp = query.execute()
        return {r.get("application_name") for r in (resp.data or []) if r.get("application_name")}
    except:
        return set()

def fetch_bugs_data(tenant_id: str, user_id: str, project: Optional[str] = None, projects: Optional[List[str]] = None):
    select_cols = ", ".join(QA_BUG_SELECT_COLUMNS)
    query = f'SELECT {select_cols} FROM "bugs" WHERE "tenant_id" = %s AND "is_deleted" IS DISTINCT FROM TRUE'
    params = [tenant_id]
    if projects:
        cleaned = []
        for p in projects:
            if p is None:
                continue
            if isinstance(p, dict):
                v = p.get("value") or p.get("label")
                if v:
                    cleaned.append(str(v).strip())
                continue
            if isinstance(p, str):
                ps = p.strip()
                if not ps:
                    continue
                if ps.startswith("{") and "value" in ps:
                    try:
                        obj = json.loads(ps)
                        if isinstance(obj, dict):
                            v = obj.get("value") or obj.get("label")
                            if v:
                                cleaned.append(str(v).strip())
                                continue
                    except Exception:
                        pass
                cleaned.append(ps)
                continue
            cleaned.append(str(p).strip())
        cleaned = [c for c in cleaned if c]
        if cleaned:
            placeholders = ",".join(["%s"] * len(cleaned))
            query += f' AND (COALESCE("Project", \'\') IN ({placeholders}) OR COALESCE("Product", \'\') IN ({placeholders}))'
            params.extend(cleaned)
            params.extend(cleaned)
    elif project and project != "All":
        query += ' AND (COALESCE("Project", \'\') = %s OR COALESCE("Product", \'\') = %s)'
        params.append(project)
        params.append(project)
        
    rows = execute_query(query, tuple(params), fetch_all=True) or []
    
    # Apply RBAC filtering
    accessible_projects = get_user_accessible_project_names(user_id, tenant_id)
    if accessible_projects is not None:
        # Normalize accessible projects (strip spaces)
        accessible_projects_clean = {str(p).strip() for p in accessible_projects if p}
        rows = [
            r for r in rows 
            if (r.get("Project") and str(r.get("Project")).strip() in accessible_projects_clean) or 
               (r.get("Product") and str(r.get("Product")).strip() in accessible_projects_clean)
        ]

    user_map = get_user_map(tenant_id)
    # Create case-insensitive map for safety
    user_map_lower = {k.lower(): v for k, v in user_map.items()}
    
    for row in rows:
        if not row.get("Project") and row.get("Product"):
            row["Project"] = row.get("Product")

        # Resolve Assignee
        assignee_val = row.get("Assignee")
        if assignee_val:
             val_str = str(assignee_val).strip().lower()
             u_info = user_map_lower.get(val_str)
             if u_info:
                 row["Assignee Real Name"] = u_info.get("name")
                 row["Assignee Email"] = u_info.get("email")
        
        # Resolve Reporter
        reporter_val = row.get("Reporter")
        if reporter_val:
             val_str = str(reporter_val).strip().lower()
             u_info = user_map_lower.get(val_str)
             if u_info:
                 row["Reporter Real Name"] = u_info.get("name")
                 row["Reporter Email"] = u_info.get("email")

        # Resolve Project Owner
        owner_val = row.get("Project Owner")
        if owner_val:
             val_str = str(owner_val).strip().lower()
             u_info = user_map_lower.get(val_str)
             if u_info:
                 row["Project Owner Name"] = u_info.get("name")
                 row["Project Owner Email"] = u_info.get("email")
    
    return rows

@router.get("/projects")
@require_permission("qa_dashboard_retrieve")
async def get_projects(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        tenant_id = user.get("tenant_id") or tenant_id

        
        query = local_db.table("projects").select("application_name").order("application_name")
        if tenant_id:
            query = query.eq("tenant_id", tenant_id)
            
        accessible_projects = get_user_accessible_project_names(user_id, tenant_id)
        if accessible_projects is not None:
            if not accessible_projects:
                return {"data": [], "error": None}
            query = query.in_("application_name", accessible_projects)
            
        resp = query.execute()
        if getattr(resp, "error", None):
            raise HTTPException(status_code=400, detail=str(resp.error))

        names = set()
        for row in resp.data or []:
            name = row.get("application_name")
            if name:
                names.add(str(name).strip())
                
        # In case local proxy doesn't fully filter, double check
        if accessible_projects is not None:
            names = {n for n in names if n in accessible_projects}

        return {"data": sorted(list(names)), "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/projects", "get_projects", return_dict=True)

@router.get("/stats")
async def get_stats(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        tenant_id = user.get("tenant_id") or tenant_id
        
        # Fetch all project bugs first, then filter in memory for complex logic
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        # Parse filter dates
        start = parse_date(start_date)
        end = parse_date(end_date)
        
        processed_total = 0
        processed_open = 0
        processed_passed = 0
        processed_others = 0
        
        today = date.today()
        
        for b in bugs:
            status = str(b.get("Status")).upper()
            created_at = parse_date(b.get("created_at"))
            updated_at = parse_date(b.get("updated_at"))
            deadline_str = b.get("Deadline")
            
            # --- Filter Logic ---
            # Total Bugs: Based on Created Date
            in_created_range = True
            if start and created_at and created_at < start: in_created_range = False
            if end and created_at and created_at > end: in_created_range = False
            
            # QA Passed: Based on Updated Date (approx for "Passed Date")
            in_updated_range = True
            if start and updated_at and updated_at < start: in_updated_range = False
            if end and updated_at and updated_at > end: in_updated_range = False
            
            # 1. Total Bugs (in date range)
            if in_created_range:
                processed_total += 1
                
            # 2. Open Bugs (in date range - Cohort view: Created in range AND currently open)
            if in_created_range and status in OPEN_STATUSES:
                processed_open += 1
                
            # 3. QA Passed (Activity view: Passed in date range)
            if in_updated_range and status in QA_PASSED_STATUSES:
                processed_passed += 1
                
            # 4. Others Bugs (Not Open AND Not QA Passed)
            # Includes null/empty status.
            if in_created_range and status not in OPEN_STATUSES and status not in QA_PASSED_STATUSES:
                processed_others += 1

        return {
            "data": {
                "total": processed_total,
                "open": processed_open,
                "qa_passed": processed_passed, # Renamed from 'closed'
                "others": processed_others
            },
            "error": None
        }
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/stats", "get_stats", return_dict=True)

@router.get("/charts/severity")
@require_permission("qa_dashboard_retrieve")
async def get_severity_chart(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        tenant_id = user.get("tenant_id") or tenant_id
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        start = parse_date(start_date)
        end = parse_date(end_date)
        
        # Aggregate
        severity_counts = {}
        
        for b in bugs:
            # Date Filter (Created At)
            created_at = parse_date(b.get("created_at"))
            if start and created_at and created_at < start: continue
            if end and created_at and created_at > end: continue
            
            severity = b.get("Severity") or "Unknown"
            status = str(b.get("Status")).upper()
            is_open = status in OPEN_STATUSES
            
            if severity not in severity_counts:
                severity_counts[severity] = {"severity": severity, "open": 0, "qa_passed": 0, "total": 0}
            
            severity_counts[severity]["total"] += 1
            if is_open:
                severity_counts[severity]["open"] += 1
            elif status in QA_PASSED_STATUSES:
                severity_counts[severity]["qa_passed"] += 1
                
        # Define order
        order = ["Blocker", "Critical", "Major", "Minor", "Low"]
        result = []
        
        # Add ordered items first
        for s in order:
            if s in severity_counts:
                result.append(severity_counts[s])
                del severity_counts[s]
        
        # Add remaining
        for s in severity_counts:
             result.append(severity_counts[s])
             
        return {"data": result, "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/charts/severity", "get_severity_chart", return_dict=True)

@router.get("/charts/status")
@require_permission("qa_dashboard_retrieve")
async def get_status_chart(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        tenant_id = user.get("tenant_id") or tenant_id
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        start = parse_date(start_date)
        end = parse_date(end_date)
        
        status_counts = {}
        for b in bugs:
            # Date Filter (Created At)
            created_at = parse_date(b.get("created_at"))
            if start and created_at and created_at < start: continue
            if end and created_at and created_at > end: continue

            status = normalize_status(b.get("Status")) or "UNKNOWN"
            status_counts[status] = status_counts.get(status, 0) + 1
            
        result = [{"name": k, "value": v} for k, v in status_counts.items()]
        # Sort by value desc
        result.sort(key=lambda x: x["value"], reverse=True)
        
        return {"data": result, "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/charts/status", "get_status_chart", return_dict=True)

@router.get("/pivot/status-defect-type")
@require_permission("qa_dashboard_retrieve")
async def pivot_status_defect_type(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        tenant_id = user.get("tenant_id") or tenant_id
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)

        start = parse_date(start_date)
        end = parse_date(end_date)

        # Build dynamic columns from defect types (no hardcoding)
        defect_types: set = set()
        status_rows: Dict[str, Dict[str, Any]] = {}

        for b in bugs:
            c_at = parse_date(b.get("created_at"))
            if start and c_at and c_at < start: 
                continue
            if end and c_at and c_at > end: 
                continue

            status = normalize_status(b.get("Status")) or "UNKNOWN"
            defect = b.get("Defect type") or b.get("Defect Type") or "Unknown"
            defect = str(defect).strip() or "Unknown"
            status = str(status).strip() or "UNKNOWN"

            defect_types.add(defect)
            if status not in status_rows:
                status_rows[status] = {"name": status, "total": 0}
            status_rows[status][defect] = (status_rows[status].get(defect, 0) + 1)
            status_rows[status]["total"] += 1

        columns = sorted(list(defect_types))
        rows = sorted(list(status_rows.values()), key=lambda r: str(r["name"]))

        # Add grand total row
        if rows:
            grand = {"name": "Total", "total": 0}
            for d in columns:
                grand[d] = 0
            for r in rows:
                grand["total"] += int(r.get("total", 0))
                for d in columns:
                    grand[d] += int(r.get(d, 0))
            rows.append(grand)

        return {"data": {"columns": columns, "rows": rows}, "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/pivot/status-defect-type", "pivot_status_defect_type", return_dict=True)
@router.get("/charts/assignee")
async def get_assignee_chart(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        start = parse_date(start_date)
        end = parse_date(end_date)
        
        user_map = get_user_map(tenant_id)
        assignee_counts = {}
        
        for b in bugs:
            # Date Filter (Created At)
            created_at = parse_date(b.get("created_at"))
            if start and created_at and created_at < start: continue
            if end and created_at and created_at > end: continue

            status = str(b.get("Status")).upper()
            # Only count open bugs
            if status not in OPEN_STATUSES:
                continue
                
            assignee_val = b.get("Assignee")
            assignee_real_name = b.get("Assignee Real Name")
            
            display_val = "Unassigned"
            
            if assignee_val:
                # Check if it's in user_map (assignee_val is likely UUID string)
                u_info = user_map.get(str(assignee_val))
                if u_info and u_info.get("email"):
                    display_val = u_info.get("email")
                elif "@" in str(assignee_val):
                     # It's already an email
                     display_val = str(assignee_val)
                elif assignee_real_name:
                     # Fallback to name if we can't find email
                     display_val = assignee_real_name
                else:
                     # Fallback to ID/Value if nothing else
                     display_val = str(assignee_val)
            elif assignee_real_name:
                display_val = assignee_real_name
                
            assignee_counts[display_val] = assignee_counts.get(display_val, 0) + 1
            
        result = [{"name": k, "value": v} for k, v in assignee_counts.items()]
        result.sort(key=lambda x: x["value"], reverse=True)
        
        # Top 10
        return {"data": result[:10], "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/charts/assignee", "get_assignee_chart", return_dict=True)

@router.get("/charts/ageing")
async def get_ageing_chart(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        start = parse_date(start_date)
        end = parse_date(end_date)
        
        buckets = {
            "0-7 days": 0,
            "8-30 days": 0,
            "31-60 days": 0,
            "60+ days": 0
        }
        
        for b in bugs:
            # Date Filter (Created At)
            created_at = parse_date(b.get("created_at"))
            
            # AGEING CHART: Ignore global date filters. 
            # if start and created_at and created_at < start: continue
            # if end and created_at and created_at > end: continue

            status = str(b.get("Status")).upper()
            
            # Only count open bugs
            if status not in OPEN_STATUSES:
                continue
                
            # Calculate age dynamically from created_at
            if not created_at:
                age_days = 0
            else:
                age_days = (date.today() - created_at).days
            
            try:
                if age_days <= 7:
                    buckets["0-7 days"] += 1
                elif age_days <= 30:
                    buckets["8-30 days"] += 1
                elif age_days <= 60:
                    buckets["31-60 days"] += 1
                else:
                    buckets["60+ days"] += 1
            except (ValueError, TypeError):
                pass

        result = [{"name": k, "value": v} for k, v in buckets.items()]
        return {"data": result, "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/charts/ageing", "get_ageing_chart", return_dict=True)

@router.get("/charts/priority")
@require_permission("qa_dashboard_retrieve")
async def get_priority_chart(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        start = parse_date(start_date)
        end = parse_date(end_date)
        
        priority_counts = {}
        for b in bugs:
            # Date Filter (Created At)
            created_at = parse_date(b.get("created_at"))
            if start and created_at and created_at < start: continue
            if end and created_at and created_at > end: continue
            p = b.get("Priority") or "Unknown"
            priority_counts[p] = priority_counts.get(p, 0) + 1
            
        result = [{"name": k, "value": v} for k, v in priority_counts.items()]
        result.sort(key=lambda x: x["value"], reverse=True)
        
        return {"data": result, "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/charts/priority", "get_priority_chart", return_dict=True)

@router.get("/charts/trend")
@require_permission("qa_dashboard_retrieve")
async def get_trend_chart(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        start = parse_date(start_date)
        end = parse_date(end_date)
        
        if not start:
             start = date.today() - timedelta(days=30)
        if not end:
             end = date.today()
             
        # Initialize dates
        date_map = {}
        curr = start
        while curr <= end:
            d_str = curr.strftime("%Y-%m-%d")
            date_map[d_str] = {
                "date": d_str,
                "created": 0,
                "qa_passed": 0,
                "open_cumulative": 0  # To be calculated
            }
            curr += timedelta(days=1)
            
        # 1. Fill Created and QA Passed counts
        for b in bugs:
            c_at = parse_date(b.get("created_at"))
            u_at = parse_date(b.get("updated_at"))
            status = str(b.get("Status")).upper()
            
            # Created count
            if c_at:
                c_str = c_at.strftime("%Y-%m-%d")
                if c_str in date_map:
                    date_map[c_str]["created"] += 1
            
            # QA Passed count (Activity based on updated_at)
            if u_at and status in QA_PASSED_STATUSES:
                u_str = u_at.strftime("%Y-%m-%d")
                if u_str in date_map:
                    date_map[u_str]["qa_passed"] += 1

        # 2. Calculate Open Trends (Snapshot approximation)
        # Convert date_map to sorted list
        chart_data = sorted(date_map.values(), key=lambda x: x['date'])
        
        # We need cumulative open for each day in the chart range.
        # This is expensive if we do it for every bug every day. 
        # Optimization: 
        #   Open[Day] = Total Created (<= Day) - Total Closed/Passed (<= Day)
        #   Provided "Closed/Passed" status is final. 
        #   "QA Passed" is one closed state. What about others like DONE, RESOLVED?
        #   User said "OPEN" includes RESOLVED. So primarily we subtract QA PASSED (if final).
        #   Wait, if status is RESOLVED it is considered "Open Bugs" bucket.
        #   So meaningful exit status is ONLY "QA PASSED" for the purpose of "Open Bugs" chart?
        #   User: "QA Passed = status in ['QA PASSED']", "Open Bugs = all statuses mapped to OPEN list"
        #   So strictly, if I am not in Open List, I am Closed?
        #   Wait, are there other statuses not in either list?
        #   OPEN_STATUSES has: OPEN, IN_PROGRESS, RESOLVED, READY FOR DEPLOYMENT, READY FOR QA, QA IN PROGRESS, QA REOPENED, + NEED CLARIFICATION...
        #   QA_PASSED_STATUSES has: QA PASSED.
        #   "Done"? "Closed"? (Old values).
        #   Let's assume anything NOT in OPEN_STATUSES is 'Inactive' or 'Passed' or 'Closed'.
        #   For the trend "Open Bugs", we want count of bugs currently in OPEN_STATUSES on that day.
        
        #   Algorithm:
        #   For each day D in range:
        #     Count bugs where:
        #       created_at <= D AND
        #       ( NOT (status == QA PASSED AND updated_at <= D) )
        #       AND ( NOT (status == CLOSED AND updated_at <= D) ) # If any other closed states exist
        #       Actually, simpler:
        #       Open[D] = Created <= D  MINUS  (QA Passed <= D)
        #       (Ignoring bugs that were passed and then reopened... wait, if reopened, updated_at changes?
        #        If I only use updated_at, I lose history. This is the "Best Effort".
        #        Using "Created <= D MINUS QaPassed <= D" is a valid approximation.)
        
        dates = sorted(list(date_map.keys()))
        for d_str in dates:
            d_date = datetime.strptime(d_str, "%Y-%m-%d").date()
            
            # Cumulative Created
            cum_created = sum(1 for b in bugs if parse_date(b.get("created_at")) and parse_date(b.get("created_at")) <= d_date)
            # Cumulative Passed (using updated_at as proxy for pass date)
            cum_passed = sum(1 for b in bugs if parse_date(b.get("updated_at")) and parse_date(b.get("updated_at")) <= d_date and str(b.get("Status")).upper() in QA_PASSED_STATUSES)
            
            # Resulting Open
            date_map[d_str]["open_cumulative"] = max(0, cum_created - cum_passed)

        return {"data": chart_data, "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/charts/trend", "get_trend_chart", return_dict=True)


@router.get("/snapshots/trend")
@require_permission("qa_dashboard_retrieve")
async def get_snapshot_trend(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None),
):
    """Daily trend from qa_bug_snapshot_daily (login-generated snapshots)."""
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")

        start = parse_date(start_date)
        end = parse_date(end_date)
        if not end:
            end = date.today()
        if not start:
            start = end - timedelta(days=30)

        project_filter: List[str] = []
        if projects:
            project_filter = [str(p).strip() for p in projects if p and str(p).strip()]
        elif project and project != "All":
            project_filter = [str(project).strip()]

        accessible = get_user_accessible_project_names(user_id, tenant_id)
        if accessible is not None:
            accessible_clean = {str(p).strip() for p in accessible if p}
            if project_filter:
                project_filter = [p for p in project_filter if p in accessible_clean]
            else:
                project_filter = list(accessible_clean)

        open_placeholders = ",".join(["%s"] * len(SNAPSHOT_OPEN_SUPER_STATUSES))
        closed_placeholders = ",".join(["%s"] * len(SNAPSHOT_CLOSED_SUPER_STATUSES))

        params: List[Any] = [
            *SNAPSHOT_OPEN_SUPER_STATUSES,
            *SNAPSHOT_CLOSED_SUPER_STATUSES,
            *SNAPSHOT_OPEN_SUPER_STATUSES,
            *SNAPSHOT_OPEN_SUPER_STATUSES,
            *SNAPSHOT_CLOSED_SUPER_STATUSES,
            *SNAPSHOT_CLOSED_SUPER_STATUSES,
            *SNAPSHOT_CLOSED_SUPER_STATUSES,
            *SNAPSHOT_CLOSED_SUPER_STATUSES,
            start,
            end,
        ]

        query = f"""
            SELECT
                snapshot_date::text AS date,
                COALESCE(SUM(total_bugs) FILTER (
                    WHERE super_status IN ({open_placeholders})
                ), 0) AS open_bugs,
                COALESCE(SUM(total_bugs) FILTER (
                    WHERE super_status IN ({closed_placeholders})
                ), 0) AS closed_bugs,
                COALESCE(SUM(total_bugs), 0) AS total_bugs,
                COALESCE(SUM(total_bugs) FILTER (
                    WHERE super_status IN ({open_placeholders})
                      AND UPPER(TRIM(priority)) IN (
                          'BLOCKER', 'HIGH', 'P1', 'P2', 'CRITICAL', 'MAJOR'
                      )
                ), 0) AS blocker_high,
                COALESCE(SUM(total_bugs) FILTER (
                    WHERE super_status IN ({open_placeholders})
                      AND age_bucket = '30+ DAYS'
                ), 0) AS sla_breached,
                COALESCE(SUM(total_bugs) FILTER (
                    WHERE super_status IN ({closed_placeholders})
                      AND UPPER(TRIM(priority)) IN ('BLOCKER', 'P1', 'CRITICAL')
                ), 0) AS closed_blocker,
                COALESCE(SUM(total_bugs) FILTER (
                    WHERE super_status IN ({closed_placeholders})
                      AND UPPER(TRIM(priority)) IN ('HIGH', 'P2', 'MAJOR')
                ), 0) AS closed_high,
                COALESCE(SUM(total_bugs) FILTER (
                    WHERE super_status IN ({closed_placeholders})
                      AND UPPER(TRIM(priority)) IN ('MEDIUM', 'P3', 'MINOR')
                ), 0) AS closed_medium,
                COALESCE(SUM(total_bugs) FILTER (
                    WHERE super_status IN ({closed_placeholders})
                      AND UPPER(TRIM(priority)) IN ('LOW', 'P4')
                ), 0) AS closed_low
            FROM qa_bug_snapshot_daily
            WHERE snapshot_date >= %s
              AND snapshot_date <= %s
        """

        if project_filter:
            placeholders = ",".join(["%s"] * len(project_filter))
            query += f" AND LOWER(TRIM(project_name)) IN ({placeholders})"
            params.extend([p.strip().lower() for p in project_filter])

        query += " GROUP BY snapshot_date ORDER BY snapshot_date"

        rows = execute_query(query, tuple(params), fetch_all=True) or []
        row_by_date = {r["date"]: r for r in rows}
        snapshot_total = sum(int(r.get("total_bugs") or 0) for r in rows)

        chart_data = []
        curr = start
        while curr <= end:
            d_str = curr.strftime("%Y-%m-%d")
            row = row_by_date.get(d_str, {})
            chart_data.append(_snapshot_metrics_to_chart_point(d_str, row))
            curr += timedelta(days=1)

        # No snapshot history: fill today from live bugs so the trend chart is not blank
        if snapshot_total == 0:
            tenant_id = user.get("tenant_id") or tenant_id
            bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
            live_metrics = _aggregate_bugs_snapshot_metrics(bugs)
            today_str = date.today().strftime("%Y-%m-%d")
            for point in chart_data:
                if point["date"] == today_str:
                    point.update(_snapshot_metrics_to_chart_point(today_str, live_metrics))
                    break

        return {"data": chart_data, "error": None}
    except Exception as e:
        return handle_endpoint_error(
            e, "/qa-dashboard/snapshots/trend", "get_snapshot_trend", return_dict=True
        )


@router.post("/snapshots/generate")
@require_permission("qa_dashboard_retrieve")
async def trigger_daily_snapshot(
    force: bool = Query(False, description="Delete today's snapshot rows and regenerate"),
    Authorization: Optional[str] = Header(default=None),
):
    """Manually run daily bug snapshot (ops / recovery after fixing DB function)."""
    try:
        auth_guard(Authorization)
        from services.qa_bug_snapshot_service import (
            force_regenerate_today_snapshot,
            generate_daily_bug_snapshot,
            today_snapshot_row_count,
        )

        if force:
            row_count = force_regenerate_today_snapshot()
        else:
            generate_daily_bug_snapshot()
            row_count = today_snapshot_row_count()

        return {
            "data": {"snapshot_date": date.today().isoformat(), "rows": row_count, "forced": force},
            "error": None,
        }
    except Exception as e:
        return handle_endpoint_error(
            e, "/qa-dashboard/snapshots/generate", "trigger_daily_snapshot", return_dict=True
        )


@router.get("/project-health")
async def get_project_health(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        
        # Get all bugs
        bugs = fetch_bugs_data(tenant_id, user_id, None) # All projects
        
        # Aggregation per project
        projects = {}
        today = date.today()
        
        for b in bugs:
            p_name = b.get("Project")
            if not p_name: continue
            
            if p_name not in projects:
                projects[p_name] = {
                    "project": p_name,
                    "open_critical": 0,
                    "open_major": 0,
                    "sla_breaches": 0,
                    "reopen_count": 0, # Assuming we can track this? field 'Reopen Count' or similar?
                    "risk_score": 0
                }
            
            status = str(b.get("Status")).upper()
            if status not in OPEN_STATUSES:
                continue
                
            priority = str(b.get("Priority") or "").upper()
            
            # Open Critical / Major
            if "CRITICAL" in priority or "BLOCKER" in priority:
                projects[p_name]["open_critical"] += 1
            if "MAJOR" in priority:
                 projects[p_name]["open_major"] += 1
                 
            # Reopen Count (If available field)
            # Assuming 'Reopen Count' exists based on request "Reopen Count * 1"
            # If not, ignore or use 0
            if b.get("Reopen Count"):
                try:
                    projects[p_name]["reopen_count"] += int(b.get("Reopen Count"))
                except:
                    pass
            
            # SLA Breach
            # Critical open bug age > 1 day
            # Major open bug age > 3 days
            bug_age = 0
            if b.get("Bug Age (in days)"):
                 try:
                     bug_age = int(b.get("Bug Age (in days)"))
                 except:
                     pass
            elif b.get("created_at"):
                 try:
                     c_date = parse_date(b.get("created_at"))
                     if c_date:
                         bug_age = (today - c_date).days
                 except:
                     pass
            
            if ("CRITICAL" in priority or "BLOCKER" in priority) and bug_age > 1:
                projects[p_name]["sla_breaches"] += 1
            elif "MAJOR" in priority and bug_age > 3:
                projects[p_name]["sla_breaches"] += 1

        # Calculate Score
        # Open Critical * 5
        # Open Major * 3
        # SLA Breaches * 2
        # Reopen Count * 1
        
        result = []
        for p_data in projects.values():
            score = (p_data["open_critical"] * 5) + \
                    (p_data["open_major"] * 3) + \
                    (p_data["sla_breaches"] * 2) + \
                    (p_data["reopen_count"] * 1)
            
            p_data["risk_score"] = score
            
            if score < 20:
                p_data["risk_tag"] = "Healthy" # Green
            elif score <= 49:
                p_data["risk_tag"] = "At Risk" # Amber
            else:
                p_data["risk_tag"] = "Critical" # Red
                
            result.append(p_data)
            
        result.sort(key=lambda x: x["risk_score"], reverse=True)
        return {"data": result, "error": None}
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/project-health", "get_project_health", return_dict=True)

@router.get("/bugs")
async def get_bug_list(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    bucket: Optional[str] = Query(None), # 'open', 'qa_passed', 'overdue', 'total'
    search: Optional[str] = Query(None),
    page: int = Query(1),
    limit: int = Query(20),
    sort_by: Optional[str] = Query("created_at"),
    sort_desc: bool = Query(True),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        tenant_id = user.get("tenant_id") or tenant_id
        
        # In-memory filtering again because of complex bucket logic and status mapping
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        start = parse_date(start_date)
        end = parse_date(end_date)
        today = date.today()
        
        filtered_bugs = []
        
        for b in bugs:
            status = str(b.get("Status")).upper()
            created_at = parse_date(b.get("created_at"))
            updated_at = parse_date(b.get("updated_at"))
            deadline_str = b.get("Deadline")
            
            # 1. Date Filter
            # For 'qa_passed' bucket, we filter by updated_at?
            # Or do we respect the global dashboard logic where Date Filter applies to 'created_at' usually?
            # User said: "Drill-down: Clicking 'QA Passed' should show a table.. with filters applied"
            # In get_stats, 'qa_passed' used updated_at filter. 'open' used created_at.
            # So we should replicate that logic based on bucket.
            
            in_range = True
            
            if bucket == "qa_passed":
                # Use updated_at
                d = updated_at
            else:
                # Use created_at for total, open, others
                d = created_at
            
            if start and d and d < start: in_range = False
            if end and d and d > end: in_range = False
            
            if not in_range:
                continue
                
            # 2. Bucket Filter
            if bucket == "open":
                if status not in OPEN_STATUSES: continue
            elif bucket == "qa_passed":
                if status not in QA_PASSED_STATUSES: continue
            elif bucket == "others":
                # Not in Open AND Not in QA Passed
                 if status in OPEN_STATUSES or status in QA_PASSED_STATUSES: continue
            
            # 3. Search
            if search:
                s = search.lower()
                # Search in ID, Title, Project, Assignee
                found = False
                if s in str(b.get("Bug ID") or "").lower(): found = True
                elif s in str(b.get("Title") or "").lower(): found = True
                elif s in str(b.get("Project") or "").lower(): found = True
                elif s in str(b.get("Assignee") or "").lower(): found = True
                
                if not found: continue
            
            # Calculate Age
            age = 0
            if created_at:
                age = (today - created_at).days
            b["Bug Age (in days)"] = age
            
            filtered_bugs.append(b)
            
        # Sort
        def get_sort_key(item):
            k = sort_by
            # Helper for date sorting
            if k == "created_at": return safe_sort_datetime(item.get("created_at"))
            if k == "updated_at": return safe_sort_datetime(item.get("updated_at"))
            if k == "age": 
                # convert to int
                try: return int(item.get("Bug Age (in days)") or 0)
                except: return 0
            
            # Safe fallback for strings to ensure case-insensitive string comparison
            val = item.get(k)
            if val is None: return ""
            return str(val).lower()

        filtered_bugs.sort(key=get_sort_key, reverse=sort_desc)
        
        # Pagination
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        paginated_data = filtered_bugs[start_idx:end_idx]
        
        return {
            "data": paginated_data,
            "meta": {
                "total": len(filtered_bugs),
                "page": page,
                "limit": limit
            },
            "error": None
        }
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/bugs", "get_bug_list", return_dict=True)

@router.get("/export")
async def export_dashboard(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    projects: Optional[List[str]] = Query(None),
    bucket: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        user = auth_guard(Authorization)
        user_id = user.get("user_id") or user.get("id")
        tenant_id = user.get("tenant_id") or tenant_id
        
        # 1. Fetch Data
        bugs = fetch_bugs_data(tenant_id, user_id, project, projects)
        
        # 2. Filter Data (Reusing get_bug_list logic)
        filtered_bugs = []
        today = date.today()
        
        for b in bugs:
             status = str(b.get("Status")).upper()
             created_at = parse_date(b.get("created_at"))
             updated_at = parse_date(b.get("updated_at"))
             deadline_str = b.get("Deadline")
             
             # Filter by Bucket
             if bucket:
                 if bucket == "total":
                     pass # No filter
                 elif bucket == "open":
                     if status not in OPEN_STATUSES: continue
                 elif bucket == "qa_passed":
                     if status not in QA_PASSED_STATUSES: continue
                 elif bucket == "others":
                     if status in OPEN_STATUSES or status in QA_PASSED_STATUSES: continue
                 elif bucket == "overdue":
                     if status not in OPEN_STATUSES: continue
                     # Check deadline
                     is_overdue = False
                     if deadline_str:
                         try:
                             d_date = parse_date(deadline_str)
                             if d_date and d_date < today:
                                 is_overdue = True
                         except: pass
                     if not is_overdue: continue
             
             # Search Filter
             if search:
                 s = search.lower()
                 text = (
                     str(b.get("Bug ID") or "") + " " +
                     str(b.get("Summary") or "") + " " +
                     str(b.get("Title") or "")
                 ).lower()
                 if s not in text:
                     continue
            
             filtered_bugs.append(b)

        # 3. Sort by Created At Desc
        def get_sort_key(item):
             return safe_sort_datetime(item.get("created_at"))
        filtered_bugs.sort(key=get_sort_key, reverse=True)
        
        # 4. Create Excel
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Headers
        headers = [
            "Bug ID", "Project", "Title", "Status", "Severity", "Priority", 
            "Assignee", "Reporter", "Created At", "Updated At", "Bug Age (Days)"
        ]
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            
        # Rows
        for row_idx, b in enumerate(filtered_bugs, 2):
            c_at = parse_date(b.get("created_at"))
            u_at = parse_date(b.get("updated_at"))
            age = 0
            if c_at:
                age = (today - c_at).days
            
            ws.cell(row=row_idx, column=1, value=b.get("Bug ID"))
            ws.cell(row=row_idx, column=2, value=b.get("Project"))
            ws.cell(row=row_idx, column=3, value=b.get("Summary") or b.get("Title")) 
            ws.cell(row=row_idx, column=4, value=b.get("Status"))
            ws.cell(row=row_idx, column=5, value=b.get("Severity"))
            ws.cell(row=row_idx, column=6, value=b.get("Priority"))
            ws.cell(row=row_idx, column=7, value=b.get("Assignee Real Name") or b.get("Assignee"))
            ws.cell(row=row_idx, column=8, value=b.get("Reporter Real Name") or b.get("Reporter"))
            ws.cell(row=row_idx, column=9, value=str(c_at) if c_at else "")
            ws.cell(row=row_idx, column=10, value=str(u_at) if u_at else "")
            ws.cell(row=row_idx, column=11, value=age)

        # Auto-width (Approximate)
        for col_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)
            
        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"qa_dashboard_export_{date.today()}.xlsx"
        
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/export", "export_dashboard", return_dict=False)

@router.get("/bugs/export")
async def export_bugs_csv(
    tenant_id: str = Query('00000000-0000-0000-0000-000000000001'),
    project: Optional[str] = Query(None),
    bucket: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None)
):
    try:
        auth = auth_guard(Authorization)
        user_id = auth.get("user_id") or auth.get("id")
        t_id = auth.get("tenant_id")
        tenant_id = t_id or tenant_id
        ensure_bug_indexes()

        accessible_projects = get_user_accessible_project_names(user_id, tenant_id)

        where_parts = ['b."tenant_id" = %s', '(b."is_deleted" IS DISTINCT FROM %s)']
        params: List[Any] = [tenant_id, True]

        if project and project != "All":
            where_parts.append('COALESCE(b."Project",b."Product") = %s')
            params.append(project)

        # Apply RBAC project scoping only when no explicit project filter is given
        if accessible_projects is not None and not (project and project != "All"):
            if accessible_projects:
                where_parts.append('( b."Project" = ANY(%s) OR b."Product" = ANY(%s) )')
                params.append(accessible_projects)
                params.append(accessible_projects)
            else:
                # No accessible projects -> prevent data leak (zero rows)
                where_parts.append('FALSE')

        if bucket:
            if bucket == "open":
                where_parts.append('UPPER(b."Status") = ANY(%s)')
                params.append(OPEN_STATUSES)
            elif bucket == "qa_passed":
                where_parts.append('UPPER(b."Status") = ANY(%s)')
                params.append(QA_PASSED_STATUSES)
            elif bucket == "others":
                where_parts.append('NOT (UPPER(b."Status") = ANY(%s) OR UPPER(b."Status") = ANY(%s))')
                params.append(OPEN_STATUSES)
                params.append(QA_PASSED_STATUSES)

        # Date filters
        date_col = 'b."Changed"' if bucket == "qa_passed" else 'b."created_at"'
        if start_date:
            where_parts.append(f'{date_col} >= %s')
            params.append(start_date)
        if end_date:
            where_parts.append(f'{date_col} <= %s')
            params.append(end_date)

        # Search filter
        if search:
            s = f"%{search.strip().lower()}%"
            where_parts.append('(LOWER(b."Bug ID") LIKE %s OR LOWER(b."Summary") LIKE %s OR LOWER(b."Description") LIKE %s OR LOWER(COALESCE(b."Project",b."Product")) LIKE %s OR LOWER(COALESCE(b."Assignee Real Name",b."Assignee")) LIKE %s)')
            params.extend([s, s, s, s, s])
        # Explicit status filter (exact match, case-insensitive)
        if status:
            where_parts.append('UPPER(b."Status") = UPPER(%s)')
            params.append(status)
        if priority:
            where_parts.append('UPPER(b."Priority") = UPPER(%s)')
            params.append(priority)

        where_clause = " WHERE " + " AND ".join(where_parts)

        select_cols = [
            'b."Bug ID" as bug_id',
            'COALESCE(b."Project",b."Product") as project',
            'b."Summary" as title',
            'b."Status" as status',
            'b."Severity" as severity',
            'b."Priority" as priority',
            'COALESCE(b."Assignee Real Name", u1.full_name, u1.email, b."Assignee") as assignee_name',
            'COALESCE(b."Reporter Real Name", u2.full_name, u2.email, b."Reporter") as reporter_name',
            'b."created_at"',
            'b."Changed" as updated_at',
            '(DATE_PART(\'day\', NOW()::timestamp - COALESCE(b."created_at"::timestamp, NOW()::timestamp)))::int as bug_age_days'
        ]
        select_clause = "SELECT " + ", ".join(select_cols)
        from_clause = ' FROM "bugs" b LEFT JOIN "users" u1 ON (u1.id::text = b."Assignee" OR LOWER(u1.email) = LOWER(b."Assignee")) LEFT JOIN "users" u2 ON (u2.id::text = b."Reporter" OR LOWER(u2.email) = LOWER(b."Reporter"))'
        order_clause = ' ORDER BY "created_at" DESC, "Changed" DESC'

        # Single-query export (no LIMIT/OFFSET) per requirements
        sql = select_clause + from_clause + where_clause + order_clause
        rows = execute_query(sql, tuple(params), fetch_all=True) or []
        print(f"[Export Bugs] Rows fetched: {len(rows)}; project={project}; bucket={bucket}; status={status}; priority={priority}; search={search}; start={start_date}; end={end_date}")

        # Build CSV
        from fastapi.responses import StreamingResponse
        def csv_iter():
            yield "Bug ID,Project,Title,Status,Severity,Priority,Assignee,Reporter,Created At,Updated At,Bug Age (Days)\n"
            for r in rows:
                def fmt_dt(v):
                    if isinstance(v, datetime):
                        return v.strftime("%Y-%m-%d %H:%M")
                    if isinstance(v, date):
                        # keep date as YYYY-MM-DD
                        return v.strftime("%Y-%m-%d")
                    s = "" if v is None else str(v)
                    return s
                row = [
                    r.get("bug_id"),
                    r.get("project"),
                    r.get("title"),
                    r.get("status"),
                    r.get("severity"),
                    r.get("priority"),
                    r.get("assignee_name"),
                    r.get("reporter_name"),
                    fmt_dt(r.get("created_at")),
                    fmt_dt(r.get("updated_at")),
                    str(r.get("bug_age_days") or "")
                ]
                out = []
                for v in row:
                    s = "" if v is None else str(v)
                    if any(c in s for c in [",", "\"", "\n"]):
                        s = "\"" + s.replace("\"", "\"\"") + "\""
                    out.append(s)
                yield ",".join(out) + "\n"
        return StreamingResponse(csv_iter(), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=bugs_export_{date.today()}.csv"})
    except Exception as e:
        return handle_endpoint_error(e, "/qa-dashboard/bugs/export", "export_bugs_csv", return_dict=False)
