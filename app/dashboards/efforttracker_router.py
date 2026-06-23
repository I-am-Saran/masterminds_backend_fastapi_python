from fastapi import APIRouter, HTTPException, Header, Query
from fastapi.responses import StreamingResponse
from typing import Optional, Dict, Any, List, cast
from io import BytesIO
from datetime import datetime, date
from services.db_service import execute_query
from services.auth_service import get_user_from_token, auth_guard
from utils.error_handler import handle_endpoint_error, handle_api_error
from services.rbac_service import require_permission
import uuid

router = APIRouter(prefix="/buildtracker", tags=["buildtracker"])


# ... (existing imports)

# Append new endpoints at the end of the file

@router.get("/stats")
@require_permission("builds_retrieve")
async def get_dashboard_stats(
    project: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None),
):
    endpoint = "/buildtracker/stats"
    try:
        auth = auth_guard(Authorization)
        user = auth.get("user") or {}
        tenant_id = user.get("tenant_id")
        clauses = ["COALESCE(b.is_deleted, false) = false", "COALESCE(p.is_deleted, false) = false"]
        params: list[Any] = []

        if tenant_id:
            clauses.append("p.tenant_id::text = %s")
            params.append(str(tenant_id))
            
        # Project filter
        if project and project.strip() and project.strip() != "All":
             try:
                 uuid.UUID(project.strip())
                 clauses.append("p.id::text = %s")
             except ValueError:
                 clauses.append("p.application_name = %s")
             params.append(project.strip())

        if status and status.strip() and status.strip() != "All":
            clauses.append("TRIM(b.signoff_status) = %s")
            params.append(status.strip())

        if start_date:
            clauses.append("b.build_arrived_date::date >= %s")
            params.append(start_date)

        if end_date:
            clauses.append("b.build_arrived_date::date <= %s")
            params.append(end_date)

        # Exclude Leave projects
        clauses.append("p.application_name NOT ILIKE '%leave%'")
        # Exclude UUID names (simplistic check)
        clauses.append("p.application_name !~ '^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'")
        
        # Transaction Type logic
        allowed_types_sql = r"""
            (
                LOWER(REGEXP_REPLACE(TRIM(COALESCE(b.transaction_type, '')), '\s+-\s+', '-', 'g')) IN (
                    'functional testing-web', 'functional testing-mobile', 'performance testing',
                    'api testing', 'cross browser testing', 'cross device testing',
                    'regression testing', 'retesting', 'ar/vr testing', 'security testing',
                    'unit testing', 'integration testing', 'user acceptance testing'
                )
                OR (
                    (b.transaction_type IS NULL OR TRIM(b.transaction_type) = '' OR TRIM(b.transaction_type) = '-' OR TRIM(b.transaction_type) = '—')
                    AND TRIM(b.signoff_status) = 'Conditional Go'
                )
            )
        """
        clauses.append(allowed_types_sql)

        where_sql = "WHERE " + " AND ".join(clauses)
        
        query = f"""
            SELECT
                COUNT(*) as total_releases,
                COUNT(DISTINCT p.id) as total_projects,
                SUM(CASE WHEN LOWER(TRIM(COALESCE(b.signoff_status, ''))) IN ('', 'unknown', 'qa inprogress', 'qa in progress', '—') THEN 1 ELSE 0 END) as pending_executions,
                SUM(CASE WHEN TRIM(b.signoff_status) = 'Go' THEN 1 ELSE 0 END) as go_count,
                SUM(CASE WHEN TRIM(b.signoff_status) = 'Conditional Go' THEN 1 ELSE 0 END) as conditional_go_count,
                SUM(CASE WHEN TRIM(b.signoff_status) = 'No-Go' THEN 1 ELSE 0 END) as no_go_count,
                SUM(CASE WHEN TRIM(b.signoff_status) = 'Build Rejected' THEN 1 ELSE 0 END) as build_rejected_count
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_sql}
        """
        
        row = execute_query(query, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {}
        
        return {
            "data": {
                "transaction_distribution": distribution,
            },
            "error": None,
        }
    except Exception as e:
        error_response, status_code = handle_endpoint_error(
            e,
            endpoint,
            "get_builds_transaction_chart",
            return_dict=True,
            include_traceback=False,
        )
        raise HTTPException(status_code=status_code, detail=error_response["error"])

@router.post("/export")
@require_permission("builds_retrieve")
async def export_buildtracker(payload: Dict[str, Any], Authorization: Optional[str] = Header(default=None)):
    endpoint = "/buildtracker/export"
    try:
        auth_data = auth_guard(Authorization)
        tenant_id = auth_data.get("tenant_id")
        
        application_name = (payload.get("application_name") or "").strip()
        application_owner = (payload.get("application_owner") or "").strip()
        application_status = (payload.get("application_status") or "").strip()

        conditions = []
        params = []
        
        if tenant_id:
            conditions.append("p.tenant_id::text = %s")
            params.append(str(tenant_id))
            
        if application_name:
            conditions.append("p.application_name = %s")
            params.append(application_name)
        if application_owner:
            conditions.append("p.project_owner = %s")
            params.append(application_owner)
        if application_status:
            conditions.append("p.status = %s")
            params.append(application_status)

        conditions.append("COALESCE(p.is_deleted, false) = false")

        where_clause = ""
        if conditions:
            where_clause = "WHERE " + " AND ".join(conditions)

        conditions_builds = conditions.copy()
        conditions_builds.append("(b.is_deleted IS NULL OR b.is_deleted = FALSE)")
        where_clause_builds = "WHERE " + " AND ".join(conditions_builds)

        projects_rows = execute_query(f"""
            SELECT
                p.application_name AS application,
                p.project_owner AS owner,
                p.qa_spoc AS qa_spoc,
                p.qa_resource_count AS qa_resource_count,
                p.arrived_date AS arrived_date,
                p.expected_closing_date AS expected_closing_date
            FROM projects p
            {where_clause}
            ORDER BY p.application_name ASC
        """, tuple(params) if params else None, fetch_all=True) or []

        builds_rows = execute_query(f"""
            SELECT
                p.application_name AS application,
                b.build_number AS build_number,
                b.signoff_status AS status,
                b.build_arrived_date AS arrived_date,
                b.build_signoff_date AS signoff_date,
                b.total_bugs AS total_bugs,
                b.open_bugs AS open_bugs
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_clause_builds}
            ORDER BY p.application_name ASC, b.build_number ASC
        """, tuple(params) if params else None, fetch_all=True) or []

        tasks_rows = execute_query(f"""
            SELECT
                p.application_name AS application,
                b.build_number AS build_number,
                t.resource_name AS resource_name,
                t.task_assigned AS task_assigned
            FROM build_tasks t
            JOIN builds b ON b.id = t.build_id
            JOIN projects p ON p.id::text = b.project_id::text
            {where_clause_builds}
            ORDER BY p.application_name ASC, b.build_number ASC, t.id ASC
        """, tuple(params) if params else None, fetch_all=True) or []

        try:
            from openpyxl import Workbook
        except Exception:
            raise HTTPException(status_code=500, detail="Excel export library not available")

        wb = Workbook()
        ws_projects = wb.active
        ws_projects.title = "Projects"
        ws_projects.append(["Application", "Owner", "QA SPOC", "QA Resource Count", "Arrived Date", "Expected Closing Date"])
        for r in projects_rows:
            ws_projects.append([
                r.get("application") or "",
                r.get("owner") or "",
                r.get("qa_spoc") or "",
                r.get("qa_resource_count") or 0,
                str(r.get("arrived_date") or "") or "",
                str(r.get("expected_closing_date") or "") or "",
            ])

        ws_builds = wb.create_sheet("Builds")
        ws_builds.append(["Application", "Build #", "Status", "Arrived Date", "Signoff Date", "Total Bugs", "Open Bugs"])
        for r in builds_rows:
            ws_builds.append([
                r.get("application") or "",
                r.get("build_number") or "",
                r.get("status") or "",
                str(r.get("arrived_date") or "") or "",
                str(r.get("signoff_date") or "") or "",
                r.get("total_bugs") if isinstance(r.get("total_bugs"), (int, float)) else (int(r.get("total_bugs") or 0) if r.get("total_bugs") is not None else 0),
                r.get("open_bugs") if isinstance(r.get("open_bugs"), (int, float)) else (int(r.get("open_bugs") or 0) if r.get("open_bugs") is not None else 0),
            ])

        ws_tasks = wb.create_sheet("Tasks")
        ws_tasks.append(["Application", "Build #", "Resource Name", "Task Assigned"])
        for r in tasks_rows:
            ws_tasks.append([
                r.get("application") or "",
                r.get("build_number") or "",
                r.get("resource_name") or "",
                r.get("task_assigned") or "",
            ])

        # Dashboard summary sheet
        total_builds_row = execute_query(f"""
            SELECT COUNT(b.id) AS total_builds
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_clause_builds}
        """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {}
        total_builds = int(total_builds_row.get("total_builds") or 0)
        signed_off_row = execute_query(f"""
            SELECT COUNT(b.id) AS signed_off
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_clause_builds} AND b.build_signoff_date IS NOT NULL
        """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {}
        signed_off = int(signed_off_row.get("signed_off") or 0)
        bugs_row = execute_query(f"""
            SELECT 
                COALESCE(SUM(b.total_bugs), 0) AS total_bugs,
                COALESCE(SUM(b.open_bugs), 0) AS open_bugs
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_clause_builds}
        """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {"total_bugs": 0, "open_bugs": 0}
        total_bugs = int(bugs_row.get("total_bugs") or 0)
        open_bugs = int(bugs_row.get("open_bugs") or 0)
        closed_bugs = max(0, total_bugs - open_bugs)
        risk_row = execute_query(f"""
            SELECT
                COALESCE(SUM(CASE WHEN br.report_type IN ('functional','automation','cybersecurity') THEN br.high_count ELSE 0 END), 0) AS high_bugs,
                COALESCE(SUM(CASE WHEN br.report_type IN ('functional','automation','cybersecurity') THEN br.medium_count ELSE 0 END), 0) AS medium_bugs
            FROM build_reports br
            JOIN builds b ON b.id = br.build_id
            JOIN projects p ON p.id::text = b.project_id::text
            {where_clause_builds}
        """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {"high_bugs": 0, "medium_bugs": 0}
        high_bugs = int(risk_row.get("high_bugs") or 0)
        medium_bugs = int(risk_row.get("medium_bugs") or 0)
        in_progress_row = execute_query(f"""
            SELECT COUNT(b.id) AS in_progress
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_clause_builds} AND (b.build_signoff_date IS NULL)
        """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {}
        in_progress = int(in_progress_row.get("in_progress") or 0)
        summary_row = execute_query(f"""
            SELECT 
                MIN(p.arrived_date) AS arrived_date,
                MIN(p.expected_closing_date) AS expected_closing_date,
                (SELECT qa_spoc FROM projects pp 
                 {"WHERE " + " AND ".join(conditions) if conditions else ""}
                 ORDER BY qa_spoc NULLS LAST LIMIT 1) AS qa_spoc
            FROM projects p
            {where_clause}
        """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {}
        arrived_date = summary_row.get("arrived_date")
        expected_closing_date = summary_row.get("expected_closing_date")
        qa_spoc = summary_row.get("qa_spoc") or ""
        status_row = execute_query(f"""
            SELECT COALESCE(b.signoff_status, 'Unknown') AS status, COUNT(*) AS cnt
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_clause_builds}
            GROUP BY COALESCE(b.signoff_status, 'Unknown')
            ORDER BY cnt DESC NULLS LAST
            LIMIT 1
        """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {"status": "Unknown"}
        overall_status = status_row.get("status") or "Unknown"
        overall_progress = int(round((signed_off / total_builds) * 100)) if total_builds > 0 else 0

        ws_dashboard = wb.create_sheet("Dashboard")
        ws_dashboard.append(["Metric", "Value"])
        ws_dashboard.append(["Overall Progress (%)", overall_progress])
        ws_dashboard.append(["Projected Launch Date", str(expected_closing_date or "")])
        ws_dashboard.append(["Open Bugs %", (int(round((open_bugs / total_bugs) * 100)) if total_bugs > 0 else 0)])
        ws_dashboard.append(["High Bugs", high_bugs])
        ws_dashboard.append(["Medium Bugs", medium_bugs])
        ws_dashboard.append(["Total Bugs", total_bugs])
        ws_dashboard.append(["Open Bugs", open_bugs])
        ws_dashboard.append(["Closed Bugs", closed_bugs])
        ws_dashboard.append(["Total Builds", total_builds])
        ws_dashboard.append(["In Progress Builds", in_progress])
        ws_dashboard.append(["Arrived Date (min)", str(arrived_date or "")])
        ws_dashboard.append(["Expected Closing Date (min)", str(expected_closing_date or "")])
        ws_dashboard.append(["QA SPOC", qa_spoc])
        ws_dashboard.append(["Overall Status (top)", overall_status])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"buildtracker_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except HTTPException:
        raise
    except Exception as e:
        error_response, status_code = handle_endpoint_error(
            e,
            endpoint,
            "export_buildtracker",
            return_dict=True,
            include_traceback=False
        )
        raise HTTPException(status_code=status_code, detail=error_response["error"])

# Additional router to serve dashboard under /buildtracker prefix via main app mounting
dashboard_router = APIRouter(tags=["Build Tracker Dashboard"])

@dashboard_router.get("/dashboard")
@require_permission("builds_retrieve")
async def get_dashboard(
    Authorization: Optional[str] = Header(default=None),
    project_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
):
    endpoint = "/buildtracker/dashboard"
    try:
        auth_data = auth_guard(Authorization)
        tenant_id = auth_data.get("tenant_id")
        
        params = []
        where_projects = "WHERE COALESCE(p.is_deleted, false) = false"
        where_builds = "WHERE COALESCE(b.is_deleted, false) = false"
        where_join = "WHERE COALESCE(p.is_deleted, false) = false AND COALESCE(b.is_deleted, false) = false"
        
        if tenant_id:
            tenant_filter = " AND p.tenant_id::text = %s"
            where_projects += tenant_filter
            # For builds, we join with projects usually, or if builds has tenant_id we use it. 
            # Assuming builds table might not have tenant_id directly or we rely on project link.
            # Safe bet: Filter projects by tenant_id in the JOIN queries.
            # In the query construction below, we join builds b JOIN projects p. So p.tenant_id check is sufficient.
            where_join += tenant_filter
            params.append(str(tenant_id))
        
        # Base bug query: status open + not deleted
        # We'll refine the join logic below
        where_bugs_base = "WHERE COALESCE(bugs.\"Status\", '') = 'Open' AND COALESCE(bugs.is_deleted, false) = false"

        # List to hold WHERE clauses for builds
        build_clauses = []
        
        if project_id is not None:
            where_projects += " AND p.id = %s"
            # For builds, we add project_id check to where_builds but not to build_clauses
            # to avoid duplicating the parameter in where_join (since where_join already has p.id)
            where_builds += " AND b.project_id = %s"
            where_join += " AND p.id = %s"
            params.append(int(project_id))
        
        # Date filtering on builds
        if start_date:
            build_clauses.append("b.build_arrived_date::date >= %s")
            params.append(start_date)
        if end_date:
            build_clauses.append("b.build_signoff_date::date <= %s")
            params.append(end_date)
            
        # Construct WHERE clause for builds
        if build_clauses:
            where_builds += " AND " + " AND ".join(build_clauses)
            where_join += " AND " + " AND ".join(build_clauses)
        
        # For bugs, we need to correlate with builds that match the criteria
        # The params for bugs query will be different (we need to re-inject them)
        # So we'll construct the query string with %s and pass params copy
        
        # Construct the EXISTS clause for bugs
        # We need a fresh set of params for the bugs query because it's a separate execution context usually,
        # but here we run queries sequentially.
        # Wait, execute_query takes a tuple. We can reuse params if the order matches.
        # But bugs query has different structure.
        
        bugs_params = []
        bugs_exists_clauses = []
        if project_id is not None:
            bugs_exists_clauses.append("b.project_id = %s")
            bugs_params.append(int(project_id))
        if start_date:
            bugs_exists_clauses.append("b.build_arrived_date::date >= %s")
            bugs_params.append(start_date)
        if end_date:
            bugs_exists_clauses.append("b.build_signoff_date::date <= %s")
            bugs_params.append(end_date)
            
        where_bugs_exists = ""
        if bugs_exists_clauses:
            where_bugs_exists = " AND ".join(bugs_exists_clauses)
            
        # Total Projects
        # If date filters are applied, we should only count projects that have matching builds?
        # Or just filter projects by arrived_date?
        # User requirement: "Start Date (Arrived Date)". 
        # If we interpret strictly: filter projects where p.arrived_date >= start_date.
        # But dashboard usually implies "Activity in this period".
        # Let's try to filter projects that have *any* build matching the criteria.
        # This aligns with "filtered data from Task Tracker".
        
        projects_params = []
        projects_join_clauses = []
        if project_id is not None:
            projects_join_clauses.append("p.id = %s")
            projects_params.append(int(project_id))
        
        # We need to filter projects based on builds if date range is present
        # OR if we strictly follow "Start Date (Arrived Date)" mapping to project's arrived date?
        # Given the context of "Task Tracker", it's likely about builds.
        # Let's use EXISTS on builds for projects too.
        
        projects_where = "WHERE COALESCE(p.is_deleted, false) = false"
        if projects_join_clauses:
            projects_where += " AND " + " AND ".join(projects_join_clauses)

        # If date filters exist, add EXISTS clause for builds
        if start_date or end_date:
             projects_where += " AND EXISTS (SELECT 1 FROM builds b WHERE b.project_id = p.id"
             if start_date:
                 projects_where += " AND b.build_arrived_date >= %s"
                 projects_params.append(start_date)
             if end_date:
                 projects_where += " AND b.build_signoff_date <= %s"
                 projects_params.append(end_date)
             projects_where += ")"

        try:
            total_projects_row = execute_query(f"""
                SELECT COUNT(p.id) AS total_projects
                FROM projects p
                {projects_where}
            """, tuple(projects_params) if projects_params else None, fetch_one=True, fetch_all=False) or {}
            total_projects = int(total_projects_row.get("total_projects") or 0)
        except Exception:
            total_projects = 0

        try:
            # Re-using params for builds query
            # We need to reconstruct params for this specific query to be safe
            builds_params = []
            if project_id is not None:
                builds_params.append(int(project_id))
            if start_date:
                builds_params.append(start_date)
            if end_date:
                builds_params.append(end_date)
                
            total_builds_row = execute_query(f"""
                SELECT COUNT(b.id) AS total_builds
                FROM builds b
                JOIN projects p ON p.id::text = b.project_id::text
                {where_builds} AND COALESCE(p.is_deleted, false) = false
            """, tuple(builds_params) if builds_params else None, fetch_one=True, fetch_all=False) or {}
            total_builds = int(total_builds_row.get("total_builds") or 0)
        except Exception:
            total_builds = 0

        try:
            # Open bugs from bugs table
            # Filter by EXISTS in builds AND ensure project is not deleted
            bugs_exists_condition = " AND ".join(bugs_exists_clauses) if bugs_exists_clauses else ""
            
            bugs_sql = f"""
                SELECT COALESCE(COUNT(*), 0) AS open_bugs
                FROM public.bugs bugs
                {where_bugs_base}
                AND EXISTS (
                    SELECT 1 FROM builds b 
                    JOIN projects p ON p.id::text = b.project_id::text 
                    WHERE b.build_number = bugs."Build number as in 0.0.0.0" 
                    AND COALESCE(p.is_deleted, false) = false
                    {(" AND " + bugs_exists_condition) if bugs_exists_condition else ""}
                )
            """
            
            bugs_open_row = execute_query(bugs_sql, tuple(bugs_params) if bugs_params else None, fetch_one=True, fetch_all=False) or {"open_bugs": 0}
            open_bugs = int(bugs_open_row.get("open_bugs") or 0)
        except Exception:
            open_bugs = 0

        try:
            # QA Resources
            # Using where_join which includes build filters
            # Params must match where_join structure: project_id (if any) -> start -> end
            # The params list 'params' we built earlier should match where_join order
            # project_id is added first, then start, then end.
            
            qa_resources_row = execute_query(f"""
                SELECT COALESCE(COUNT(DISTINCT t.resource_name), 0) AS qa_resources
                FROM build_tasks t
                JOIN builds b ON b.id = t.build_id
                JOIN projects p ON p.id = b.project_id
                {where_join}
            """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {"qa_resources": 0}
            qa_resources = int(qa_resources_row.get("qa_resources") or 0)
        except Exception:
            qa_resources = 0
            
        try:
            # Total Hours
            total_hours_row = execute_query(f"""
                SELECT COALESCE(SUM(t.spent_hours), 0) AS total_hours
                FROM build_tasks t
                JOIN builds b ON b.id = t.build_id
                JOIN projects p ON p.id = b.project_id
                {where_join}
            """, tuple(params) if params else None, fetch_one=True, fetch_all=False) or {"total_hours": 0}
            total_hours = float(total_hours_row.get("total_hours") or 0)
        except Exception:
            total_hours = 0.0
            
        try:
            # Applications List
            # We want to list projects and their build stats filtered by date
            # If we filter builds in the JOIN/WHERE, we might lose projects that don't have builds in that range.
            # But usually a dashboard filter implies "Show me what happened".
            # If a project had no builds in that range, it shouldn't be in the list?
            # Or should it be there with 0s?
            # Let's assume we show projects that match the criteria (via where_join).
            # where_join enforces the build filters.
            
            # However, if we use JOIN builds b ... {where_join}, we only get projects with builds.
            # If we want all projects but filtered stats, we need LEFT JOIN and move filters to ON clause?
            # But simpler is to show only relevant projects.
            
            applications_rows = execute_query(f"""
                SELECT
                    p.id AS project_id,
                    p.application_name AS application_name,
                    p.application_type AS type,
                    MAX(b.build_number) AS latest_build_no,
                    MAX(b.signoff_status) AS latest_status,
                    p.arrived_date AS arrived_date
                FROM projects p
                JOIN builds b ON p.id::text = b.project_id::text
                {where_join}
                GROUP BY p.id, p.application_name, p.application_type, p.arrived_date
                ORDER BY p.application_name ASC
            """, tuple(params) if params else None, fetch_all=True) or []
        except Exception:
            applications_rows = []

        applications = []
        for app in applications_rows or []:
            builds_rows = execute_query("""
                SELECT
                    b.id AS id,
                    b.build_number AS build_number,
                    b.signoff_status AS status,
                    b.build_arrived_date AS arrived_date,
                    b.build_signoff_date AS signoff_date,
                    b.total_bugs AS total_bugs,
                    b.open_bugs AS open_bugs
                FROM builds b
                WHERE b.project_id = %s
                ORDER BY b.build_number ASC
            """, (app.get("project_id"),), fetch_all=True) or []
            applications.append({
                "project_id": app.get("project_id"),
                "application_name": app.get("application_name"),
                "type": app.get("type"),
                "latest_build_no": app.get("latest_build_no"),
                "latest_status": app.get("latest_status"),
                "arrived_date": app.get("arrived_date"),
                "builds": builds_rows
            })

        # Status pies (projects, builds, bugs)
        try:
            # We must use projects_where and projects_params which correctly account for
            # both project_id filtering AND date filtering (via EXISTS clause)
            # This ensures consistency with total_projects count and avoids parameter mismatch
            project_status_rows = execute_query(f"""
                SELECT COALESCE(p.status, 'Unknown') AS status, COUNT(*) AS cnt
                FROM projects p
                {projects_where}
                GROUP BY COALESCE(p.status, 'Unknown')
            """, tuple(projects_params) if projects_params else None, fetch_all=True) or []
        except Exception:
            project_status_rows = []
        try:
            build_status_rows = execute_query(f"""
                SELECT COALESCE(b.signoff_status, 'Unknown') AS status, COUNT(*) AS cnt
                FROM builds b
                JOIN projects p ON p.id::text = b.project_id::text
                {where_builds} AND COALESCE(p.is_deleted, false) = false
                GROUP BY COALESCE(b.signoff_status, 'Unknown')
            """, tuple(params) if params else None, fetch_all=True) or []
        except Exception:
            build_status_rows = []
        try:
            # Also apply project deletion check for bug status pie
            bugs_exists_condition = " AND ".join(bugs_exists_clauses) if bugs_exists_clauses else ""
            bug_status_rows = execute_query(f"""
                SELECT COALESCE(bugs."Status", 'Unknown') AS status, COUNT(*) AS cnt
                FROM public.bugs bugs
                {where_bugs_base}
                AND EXISTS (
                    SELECT 1 FROM builds b 
                    JOIN projects p ON p.id::text = b.project_id::text 
                    WHERE b.build_number = bugs."Build number as in 0.0.0.0" 
                    AND COALESCE(p.is_deleted, false) = false
                    {(" AND " + bugs_exists_condition) if bugs_exists_condition else ""}
                )
                GROUP BY COALESCE(bugs."Status", 'Unknown')
            """, tuple(bugs_params) if bugs_params else None, fetch_all=True) or []
        except Exception:
            bug_status_rows = []

        project_status_pie = [{"label": r.get("status") or "Unknown", "count": int(r.get("cnt") or 0)} for r in (project_status_rows or [])]
        build_status_pie = [{"label": r.get("status") or "Unknown", "count": int(r.get("cnt") or 0)} for r in (build_status_rows or [])]
        bug_status_pie = [{"label": r.get("status") or "Unknown", "count": int(r.get("cnt") or 0)} for r in (bug_status_rows or [])]

        print("Dashboard loaded")
        return {
            "total_projects": total_projects,
            "total_builds": total_builds,
            "open_bugs": open_bugs,
            "qa_resources": qa_resources,
            "total_hours": total_hours,
            "project_status_pie": project_status_pie,
            "build_status_pie": build_status_pie,
            "bug_status_pie": bug_status_pie,
            "applications": applications
        }
    except Exception:
        return {
            "total_projects": 0,
            "total_builds": 0,
            "open_bugs": 0,
            "qa_resources": 0,
            "total_hours": 0,
            "project_status_pie": [],
            "build_status_pie": [],
            "bug_status_pie": [],
            "applications": []
        }


@router.get("/hours/user-vs-project")
async def get_hours_user_vs_project(
    status: Optional[str] = Query(None),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    Authorization: Optional[str] = Header(default=None),
):
    endpoint = "/buildtracker/hours/user-vs-project"
    try:
        auth = auth_guard(Authorization)
        user = auth.get("user") or {}
        tenant_id = user.get("tenant_id")

        def build_where(apply_tenant: bool):
            clauses = []
            params: list[Any] = []
            
            # Filter out deleted builds
            clauses.append("(b.is_deleted IS NULL OR b.is_deleted = FALSE)")
            # Filter out deleted projects
            clauses.append("(p.is_deleted IS NULL OR p.is_deleted = FALSE)")
            
            if apply_tenant and tenant_id:
                clauses.append("((p.tenant_id::text = %s) OR p.tenant_id IS NULL)")
                params.append(str(tenant_id))
            
            if status and status.strip() and status.strip() != "All":
                s_val = status.strip()
                if s_val == "QA inprogress":
                     # Match explicit "QA inprogress" (case-insensitive) OR implicit (NULL, empty, Unknown)
                    clauses.append("(TRIM(b.signoff_status) ILIKE %s OR b.signoff_status IS NULL OR TRIM(b.signoff_status) = '' OR TRIM(b.signoff_status) = 'Unknown')")
                    params.append(s_val)
                else:
                    clauses.append("TRIM(b.signoff_status) = %s")
                    params.append(s_val)

            if start_date:
                clauses.append("bte.log_date >= %s")
                params.append(start_date)
            if end_date:
                clauses.append("bte.log_date <= %s")
                params.append(end_date)
            
            where_sql = ""
            if clauses:
                where_sql = "WHERE " + " AND ".join(clauses)
            return where_sql, params

        def query_time_entries(apply_tenant: bool):
            where_sql, params = build_where(apply_tenant)
            return execute_query(
                f"""
                SELECT
                    COALESCE(NULLIF(bte.resource_name, ''), NULLIF(bte.user_email, ''), 'Unknown') AS user_key,
                    COALESCE(p.application_name, 'Unknown') AS project_name,
                    COALESCE(SUM(bte.hours), 0) AS total_hours,
                    MAX(bte.log_date) as last_worked
                FROM build_time_entries bte
                JOIN builds b ON b.id = bte.build_id
                JOIN projects p ON p.id::text = b.project_id::text
                {where_sql}
                GROUP BY COALESCE(NULLIF(bte.resource_name, ''), NULLIF(bte.user_email, ''), 'Unknown'),
                         COALESCE(p.application_name, 'Unknown')
                """,
                tuple(params) if params else None,
                fetch_all=True,
            ) or []

        combined_rows: list[dict[str, Any]] = []

        for apply_tenant in (True, False):
            try:
                te_rows = query_time_entries(apply_tenant)
                if te_rows:
                    combined_rows.extend(te_rows)
                    break # specific match found
            except Exception:
                pass
        
        # Note: query_tasks removed as build_tasks table is empty and has schema mismatch (bigint vs uuid)

        data = []
        # Key: (user, project) -> {hours: float, last_worked: str/date}
        aggregate: dict[tuple[str, str], dict] = {}
        
        for r in combined_rows:
            user_key = r.get("user_key") or r.get("user") or "Unknown"
            project_name = r.get("project_name") or r.get("project") or "Unknown"
            raw_hours = r.get("total_hours") or r.get("hours") or 0
            row_last_worked = r.get("last_worked")
            
            try:
                hours = float(raw_hours)
            except Exception:
                try:
                    hours = float(str(raw_hours))
                except Exception:
                    hours = 0.0
            
            key = (user_key, project_name)
            if key not in aggregate:
                aggregate[key] = {"hours": 0.0, "last_worked": None}
            
            aggregate[key]["hours"] += hours
            
            # Update last_worked if newer
            curr_last = aggregate[key]["last_worked"]
            if row_last_worked:
                # Convert both to string for comparison or keep as is if comparable
                # Assuming date/datetime objects or ISO strings.
                # Simple string comparison works for ISO dates, but let's be safe if mixed types
                if curr_last is None:
                     aggregate[key]["last_worked"] = row_last_worked
                else:
                    # Prefer the greater value
                    try:
                        if str(row_last_worked) > str(curr_last):
                             aggregate[key]["last_worked"] = row_last_worked
                    except Exception:
                        pass

        for (user_key, project_name), info in aggregate.items():
            data.append(
                {
                    "user": user_key,
                    "project": project_name,
                    "hours": info["hours"],
                    "last_worked": str(info["last_worked"]) if info["last_worked"] else None
                }
            )

        return {"data": data}
    except Exception as e:
        error_response, status_code = handle_endpoint_error(
            e,
            endpoint,
            "get_hours_user_vs_project",
            return_dict=True,
            include_traceback=False,
        )
        raise HTTPException(status_code=status_code, detail=error_response["error"])


@router.get("/hours/analytics")
async def get_hours_analytics(
    status: Optional[str] = Query(None),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    Authorization: Optional[str] = Header(default=None),
):
    endpoint = "/buildtracker/hours/analytics"
    try:
        auth = auth_guard(Authorization)
        user = auth.get("user") or {}
        tenant_id = user.get("tenant_id")

        def build_where(apply_tenant: bool):
            clauses = []
            params: list[Any] = []
            
            # Filter out deleted builds
            clauses.append("(b.is_deleted IS NULL OR b.is_deleted = FALSE)")
            # Filter out deleted projects
            clauses.append("(p.is_deleted IS NULL OR p.is_deleted = FALSE)")

            if apply_tenant and tenant_id:
                clauses.append("((p.tenant_id::text = %s) OR p.tenant_id IS NULL)")
                params.append(str(tenant_id))
            
            if status and status.strip() and status.strip() != "All":
                s_val = status.strip()
                if s_val == "QA inprogress":
                    # Match explicit "QA inprogress" (case-insensitive) OR implicit (NULL, empty, Unknown)
                    clauses.append("(TRIM(b.signoff_status) ILIKE %s OR b.signoff_status IS NULL OR TRIM(b.signoff_status) = '' OR TRIM(b.signoff_status) = 'Unknown')")
                    params.append(s_val)
                else:
                    clauses.append("TRIM(b.signoff_status) = %s")
                    params.append(s_val)
            
            if start_date:
                clauses.append("bte.log_date >= %s")
                params.append(start_date)
            if end_date:
                clauses.append("bte.log_date <= %s")
                params.append(end_date)
                
            where_sql = ""
            if clauses:
                where_sql = "WHERE " + " AND ".join(clauses)
            return where_sql, params

        rows: list[dict[str, Any]] = []
        for apply_tenant in (True, False):
            where_sql, params = build_where(apply_tenant)
            try:
                rows = execute_query(
                    f"""
                    SELECT
                        COALESCE(NULLIF(bte.user_email, ''), NULLIF(bte.resource_name, ''), 'Unknown') AS user_key,
                        COALESCE(bte.log_date::date, b.build_arrived_date::date) AS work_date,
                        COALESCE(p.application_name, 'Unknown') AS project_name,
                        COALESCE(b.transaction_type, 'Unknown') AS category,
                        COALESCE(bte.hours, 0) AS hours
                    FROM build_time_entries bte
                    JOIN builds b ON b.id = bte.build_id
                    JOIN projects p ON p.id::text = b.project_id::text
                    {where_sql}
                    """,
                    tuple(params) if params else None,
                    fetch_all=True,
                ) or []
            except Exception:
                rows = []
            if rows:
                break

        by_resource: dict[str, float] = {}
        by_date: dict[str, dict[str, float]] = {}
        by_category: dict[str, float] = {}

        for r in rows or []:
            user_key = r.get("user_key") or "Unknown"
            work_date = r.get("work_date")
            category = r.get("category") or "Unknown"
            raw_hours = r.get("hours") or 0
            try:
                hours = float(raw_hours)
            except Exception:
                try:
                    hours = float(str(raw_hours))
                except Exception:
                    hours = 0.0

            by_resource[user_key] = by_resource.get(user_key, 0.0) + hours

            date_str = None
            if isinstance(work_date, datetime):
                date_str = work_date.date().isoformat()
            elif isinstance(work_date, date):
                date_str = work_date.isoformat()
            elif work_date is not None:
                date_str = str(work_date)

            if date_str:
                per_user = by_date.setdefault(date_str, {})
                per_user[user_key] = per_user.get(user_key, 0.0) + hours

            by_category[category] = by_category.get(category, 0.0) + hours

        hours_by_resource = [
            {"user": u, "hours": round(h, 2)} for u, h in by_resource.items()
        ]
        hours_by_resource.sort(key=lambda x: x["hours"], reverse=True)

        users = sorted({row["user"] for row in hours_by_resource}) if hours_by_resource else []

        hours_by_date: list[dict[str, Any]] = []
        for date_str, per_user in sorted(by_date.items()):
            row = {"date": date_str}
            for u in users:
                row[u] = round(per_user.get(u, 0.0), 2)
            hours_by_date.append(row)

        work_category = [
            {"category": c, "hours": round(h, 2)} for c, h in by_category.items()
        ]
        work_category.sort(key=lambda x: x["hours"], reverse=True)

        return {
            "data": {
                "hours_by_resource": hours_by_resource,
                "hours_by_date": hours_by_date,
                "work_category": work_category,
                "users": users,
            }
        }
    except Exception as e:
        error_response, status_code = handle_api_error(
            e, endpoint, "get_dashboard_stats", include_traceback=False
        )
        raise HTTPException(status_code=status_code, detail=error_response["error"])

@router.get("/charts/status")
async def get_status_breakdown(
    project: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None),
):
    # Reusing stats logic or just calling the stats endpoint internally?
    # Better to just query again for simplicity and independence.
    # Actually, the stats endpoint returns the counts needed for the chart.
    # The frontend chart needs: [{name: "Go", value: 10}, ...]
    # We can just format the stats data here.
    
    # ... (same setup as stats) ...
    endpoint = "/buildtracker/charts/status"
    try:
        # call the stats logic directly to avoid code duplication?
        # Or just copy paste the logic (safer to avoid coupling if logic diverges).
        # I'll copy paste for now but keep it clean.
        
        auth = auth_guard(Authorization)
        user = auth.get("user") or {}
        tenant_id = user.get("tenant_id")
        
        clauses = ["COALESCE(b.is_deleted, false) = false", "COALESCE(p.is_deleted, false) = false"]
        params = []
        
        if tenant_id:
            clauses.append("p.tenant_id::text = %s")
            params.append(str(tenant_id))
            
        if project and project.strip() and project.strip() != "All":
             try:
                 uuid.UUID(project.strip())
                 clauses.append("p.id::text = %s")
             except ValueError:
                 clauses.append("p.application_name = %s")
             params.append(project.strip())

        if status and status.strip() and status.strip() != "All":
            clauses.append("TRIM(b.signoff_status) = %s")
            params.append(status.strip())

        if start_date:
            clauses.append("b.build_arrived_date::date >= %s")
            params.append(start_date)

        if end_date:
            clauses.append("b.build_arrived_date::date <= %s")
            params.append(end_date)

        clauses.append("p.application_name NOT ILIKE '%leave%'")
        clauses.append("p.application_name !~ '^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'")
        
        allowed_types_sql = r"""
            (
                LOWER(REGEXP_REPLACE(TRIM(COALESCE(b.transaction_type, '')), '\s+-\s+', '-', 'g')) IN (
                    'functional testing-web', 'functional testing-mobile', 'performance testing',
                    'api testing', 'cross browser testing', 'cross device testing',
                    'regression testing', 'retesting', 'ar/vr testing', 'security testing',
                    'unit testing', 'integration testing', 'user acceptance testing'
                )
                OR (
                    (b.transaction_type IS NULL OR TRIM(b.transaction_type) = '' OR TRIM(b.transaction_type) = '-' OR TRIM(b.transaction_type) = '—')
                    AND TRIM(b.signoff_status) = 'Conditional Go'
                )
            )
        """
        clauses.append(allowed_types_sql)

        where_sql = "WHERE " + " AND ".join(clauses)
        
        # Group by status
        # But we need normalized status names like frontend
        # Go, Conditional Go, No-Go, Build Rejected, QA In Progress
        
        query = f"""
            SELECT
                CASE 
                    WHEN TRIM(b.signoff_status) = 'Go' THEN 'Go'
                    WHEN TRIM(b.signoff_status) = 'Conditional Go' THEN 'Conditional Go'
                    WHEN TRIM(b.signoff_status) = 'No-Go' THEN 'No-Go'
                    WHEN TRIM(b.signoff_status) = 'Build Rejected' THEN 'Build Rejected'
                    ELSE 'QA In Progress'
                END as status_group,
                COUNT(*) as count
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_sql}
            GROUP BY status_group
        """
        
        rows = execute_query(query, tuple(params) if params else None, fetch_all=True) or []
        
        # Ensure all expected statuses are present
        expected = ["Go", "Conditional Go", "No-Go", "Build Rejected", "QA In Progress"]
        result_map = {r.get("status_group"): int(r.get("count") or 0) for r in rows}
        
        data = [{"name": s, "value": result_map.get(s, 0)} for s in expected]
        
        return {"data": data}
    except Exception as e:
        error_response, status_code = handle_api_error(
            e, endpoint, "get_status_breakdown", include_traceback=False
        )
        raise HTTPException(status_code=status_code, detail=error_response["error"])

@router.get("/charts/resource-allocation")
async def get_resource_allocation(
    project: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None),
):
    endpoint = "/buildtracker/charts/resource-allocation"
    try:
        auth = auth_guard(Authorization)
        user = auth.get("user") or {}
        tenant_id = user.get("tenant_id")
        
        # Base clauses for TIME ENTRIES
        # join build_time_entries bt, builds b, projects p
        
        clauses = ["COALESCE(b.is_deleted, false) = false", "COALESCE(p.is_deleted, false) = false"]
        params = []
        
        if tenant_id:
            clauses.append("p.tenant_id::text = %s")
            params.append(str(tenant_id))
            
        if project and project.strip() and project.strip() != "All":
             try:
                 uuid.UUID(project.strip())
                 clauses.append("p.id::text = %s")
             except ValueError:
                 clauses.append("p.application_name = %s")
             params.append(project.strip())

        if status and status.strip() and status.strip() != "All":
            clauses.append("TRIM(b.signoff_status) = %s")
            params.append(status.strip())

        # Date filter applies to LOG DATE for time entries
        if start_date:
            clauses.append("bt.log_date >= %s")
            params.append(start_date)

        if end_date:
            clauses.append("bt.log_date <= %s")
            params.append(end_date)

        clauses.append("p.application_name NOT ILIKE '%leave%'")
        clauses.append("p.application_name !~ '^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'")
        
        # NO transaction type filter here (matches frontend logic)
        
        where_sql = "WHERE " + " AND ".join(clauses)
        
        query = f"""
            SELECT
                COALESCE(b.transaction_type, '—') as t_type,
                COALESCE(p.application_name, 'Unknown') as project_name,
                COALESCE(bt.resource_name, bt.user_email, 'Unknown') as resource_name,
                SUM(bt.hours) as total_hours
            FROM build_time_entries bt
            JOIN builds b ON bt.build_id = b.id
            JOIN projects p ON b.project_id::text = p.id::text
            {where_sql}
            GROUP BY b.transaction_type, p.application_name, bt.resource_name, bt.user_email
        """
        
        rows = execute_query(query, tuple(params) if params else None, fetch_all=True) or []
        
        # Transform to frontend format
        # data: [{name: "Type - Project", tType, project, resourceCount, uniqueResources (set->size), resource1: hours, resource2: hours...}]
        # resources: list of unique resource names
        
        data_map = {}
        all_resources = set()
        
        for r in rows:
            t_type = r.get("t_type") or "—"
            project_name = r.get("project_name") or "Unknown"
            resource = r.get("resource_name") or "Unknown"
            hours = float(r.get("total_hours") or 0)
            
            key = f"{t_type} - {project_name}"
            
            if key not in data_map:
                data_map[key] = {
                    "name": key,
                    "tType": t_type,
                    "project": project_name,
                    "resourceCount": 0,
                    "uniqueResources": set(),
                    "placeholder": 0
                }
            
            # Add hours to resource
            current_hours = data_map[key].get(resource, 0)
            data_map[key][resource] = current_hours + hours
            
            data_map[key]["uniqueResources"].add(resource)
            all_resources.add(resource)
            
        # Finalize counts
        result_data = []
        for key, val in data_map.items():
            val["resourceCount"] = len(val["uniqueResources"])
            del val["uniqueResources"] # remove set before returning
            result_data.append(val)
            
        return {
            "data": {
                "data": result_data,
                "resources": sorted(list(all_resources))
            }
        }
    except Exception as e:
        error_response, status_code = handle_api_error(
            e, endpoint, "get_resource_allocation", include_traceback=False
        )
        raise HTTPException(status_code=status_code, detail=error_response["error"])

@router.get("/charts/user-hours")
async def get_user_hours_chart(
    project: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None),
):
    endpoint = "/buildtracker/charts/user-hours"
    try:
        auth = auth_guard(Authorization)
        user = auth.get("user") or {}
        tenant_id = user.get("tenant_id")
        
        clauses = ["COALESCE(b.is_deleted, false) = false", "COALESCE(p.is_deleted, false) = false"]
        params = []
        
        if tenant_id:
            clauses.append("p.tenant_id::text = %s")
            params.append(str(tenant_id))
            
        if project and project.strip() and project.strip() != "All":
             try:
                 uuid.UUID(project.strip())
                 clauses.append("p.id::text = %s")
             except ValueError:
                 clauses.append("p.application_name = %s")
             params.append(project.strip())

        if status and status.strip() and status.strip() != "All":
            clauses.append("TRIM(b.signoff_status) = %s")
            params.append(status.strip())

        if start_date:
            clauses.append("bt.log_date >= %s")
            params.append(start_date)

        if end_date:
            clauses.append("bt.log_date <= %s")
            params.append(end_date)

        clauses.append("p.application_name NOT ILIKE '%leave%'")
        clauses.append("p.application_name !~ '^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'")
        
        where_sql = "WHERE " + " AND ".join(clauses)
        
        query = f"""
            SELECT
                COALESCE(bt.resource_name, bt.user_email, 'Unknown') as user_name,
                SUM(bt.hours) as total_hours
            FROM build_time_entries bt
            JOIN builds b ON bt.build_id = b.id
            JOIN projects p ON b.project_id::text = p.id::text
            {where_sql}
            GROUP BY bt.resource_name, bt.user_email
            ORDER BY total_hours DESC
        """
        
        rows = execute_query(query, tuple(params) if params else None, fetch_all=True) or []
        
        data = [
            {
                "user": r.get("user_name") or "Unknown",
                "hours": float(r.get("total_hours") or 0)
            }
            for r in rows
        ]
            
        return {"data": data}
    except Exception as e:
        error_response, status_code = handle_api_error(
            e, endpoint, "get_user_hours_chart", include_traceback=False
        )
        raise HTTPException(status_code=status_code, detail=error_response["error"])

@router.get("/charts/builds")
async def get_builds_by_transaction_type(
    project: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    Authorization: Optional[str] = Header(default=None),
):
    endpoint = "/buildtracker/charts/builds"
    try:
        auth = auth_guard(Authorization)
        user = auth.get("user") or {}
        tenant_id = user.get("tenant_id")
        
        clauses = ["COALESCE(b.is_deleted, false) = false", "COALESCE(p.is_deleted, false) = false"]
        params = []
        
        if tenant_id:
            clauses.append("p.tenant_id::text = %s")
            params.append(str(tenant_id))
            
        if project and project.strip() and project.strip() != "All":
             try:
                 uuid.UUID(project.strip())
                 clauses.append("p.id::text = %s")
             except ValueError:
                 clauses.append("p.application_name = %s")
             params.append(project.strip())

        if status and status.strip() and status.strip() != "All":
            clauses.append("TRIM(b.signoff_status) = %s")
            params.append(status.strip())

        if start_date:
            clauses.append("b.build_arrived_date::date >= %s")
            params.append(start_date)

        if end_date:
            clauses.append("b.build_arrived_date::date <= %s")
            params.append(end_date)

        clauses.append("p.application_name NOT ILIKE '%leave%'")
        clauses.append("p.application_name !~ '^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'")
        
        # Transaction Type logic (reused for filtering empty types if needed, but here we group by it)
        # We want to group by normalized type
        
        where_sql = "WHERE " + " AND ".join(clauses)
        
        query = fr"""
            SELECT
                LOWER(REGEXP_REPLACE(TRIM(COALESCE(b.transaction_type, '')), '\s+-\s+', '-', 'g')) as normalized_type,
                COUNT(*) as count
            FROM builds b
            JOIN projects p ON p.id::text = b.project_id::text
            {where_sql}
            GROUP BY normalized_type
        """
        
        rows = execute_query(query, tuple(params) if params else None, fetch_all=True) or []
        
        data = []
        for r in rows:
            t_type = r.get("normalized_type")
            if not t_type or t_type == "unknown":
                display_name = "Unknown"
            else:
                # Convert "functional testing-web" -> "Functional Testing - Web"
                # Split by hyphen if exists
                if "-" in t_type:
                    parts = t_type.split("-")
                    # Capitalize each part
                    display_name = " - ".join([p.strip().title() for p in parts])
                    # Specific fixes for common acronyms if needed (e.g. API)
                    if "Api" in display_name:
                        display_name = display_name.replace("Api", "API")
                    if "Vr" in display_name:
                         display_name = display_name.replace("Vr", "VR")
                else:
                    display_name = t_type.title()
            
            data.append({
                "name": display_name,
                "value": int(r.get("count") or 0)
            })
            
        return {"transaction_distribution": data}

    except Exception as e:
        error_response, status_code = handle_api_error(
            e, endpoint, "get_builds_by_transaction_type", include_traceback=False
        )
        raise HTTPException(status_code=status_code, detail=error_response["error"])
